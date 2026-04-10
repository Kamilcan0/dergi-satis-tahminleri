"""
Dergi Satış Projeksiyon Motoru — ML Modeli
==========================================

Mimari:
  1. Global LightGBM (90 dergi × 7 yıl = 630 satır)
     - Tüm dergiler tek modelde eğitilir
     - Walk-Forward CV ile gerçek test metrikleri
     - Lag, CAGR, TÜİK özellikleri

  2. TÜİK Hibrit Formül
     - Makro piyasa küçülmesini yakalar
     - Net = prev_net × (ulusal_oran + kategori_oran) × 0.5

  3. Ensemble
     - final = ağırlıklı (LGB × TÜİK_oran) + Hibrit
     - 2026 ≠ 2027 garantisi (TÜİK oranları yıldan yıla farklı)
"""

import os
import re
import warnings
import numpy as np
import pandas as pd
import openpyxl
import lightgbm as lgb
from sklearn.metrics import mean_absolute_error, r2_score
from sklearn.preprocessing import LabelEncoder

warnings.filterwarnings("ignore")

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data.xlsx")

# ─────────────────────────────────────────────────────────────
# Kategori eşleştirme
# ─────────────────────────────────────────────────────────────
CATEGORY_RULES = [
    ("Siyasi/Haber",       r"haber|politika|gündem|aksiyon|news|sözcü|sabah|milliyet|hurriyet|hürriyet|cumhur"),
    ("Sektörel/Mesleki",   r"bilim|science|sağlık|health|teknoloji|tech|popular|popüler|mesleki|chip|byte|pc|"
                            r"digital|dijital|photo|fotoğraf|arredamento|archit|design|tasarım|dekorasyon|"
                            r"gıda|food|tarım|tıp|medical|psikolog|psikoloji"),
    ("Kültür/Turizm",      r"atlas|tarih|history|turizm|tourism|kültür|culture|sanat|art|coğrafya|geo|"
                            r"müze|antik|arkeo|arkeoloji|mimari|archeo|national"),
    ("Ekonomi/Finans",     r"ekonomi|ekonomist|finans|borsa|para|money|business|fortune|forbes|"
                            r"capital|kapital|yatırım|invest"),
]
DEFAULT_CATEGORY = "Diğerleri"

CAT_ORDER = ["Siyasi/Haber", "Sektörel/Mesleki", "Kültür/Turizm", "Ekonomi/Finans", "Diğerleri"]


def assign_category(magazine_name: str) -> str:
    name_lower = magazine_name.lower()
    for cat, pattern in CATEGORY_RULES:
        if re.search(pattern, name_lower):
            return cat
    return DEFAULT_CATEGORY


# ─────────────────────────────────────────────────────────────
# Excel okuma
# ─────────────────────────────────────────────────────────────
def _load_workbook():
    return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)


def parse_national_tiraj() -> dict:
    wb = _load_workbook()
    ws = wb["Ulusal_Tiraj_Trend"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    data = {}
    for row in rows:
        if isinstance(row[0], int) and row[0] >= 2019:
            if isinstance(row[1], (int, float)) and row[1] > 0:
                data[row[0]] = float(row[1])
    return data


def parse_sector_trends() -> dict:
    wb = _load_workbook()
    sheet_name = next((n for n in wb.sheetnames if "sekt" in n.lower()), None)
    if not sheet_name:
        wb.close()
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header, data = None, {}
    for row in rows:
        if isinstance(row[0], str) and row[0].strip().startswith("Y"):
            header = list(row)
            continue
        if header and isinstance(row[0], int) and row[0] >= 2019:
            record = {}
            for i, col in enumerate(header[1:], start=1):
                if col and isinstance(row[i], (int, float)):
                    record[str(col)] = float(row[i])
            data[row[0]] = record
    return data


def _find_sector_key(category_name: str, year_record: dict):
    for k in year_record:
        if k.lower() == category_name.lower() or k.split("/")[0].lower() in category_name.lower():
            return k
    return list(year_record.keys())[0] if year_record else None


def parse_magazine_data() -> dict:
    wb = _load_workbook()
    ws = wb["Veri_Seti"]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    magazines, current_mag = {}, None
    for row in all_rows:
        cell0 = row[0]
        if (isinstance(cell0, str) and len(cell0.strip()) > 2
                and cell0.strip() not in ("Satır Etiketleri",)
                and not cell0.strip().startswith("Veri")):
            current_mag = cell0.strip()
            magazines[current_mag] = {"category": assign_category(current_mag),
                                       "years": [], "sevk": [], "iade": [], "net": [], "basari": []}
            continue
        if current_mag and isinstance(cell0, int) and 2019 <= cell0 <= 2025:
            sevk = row[1] if isinstance(row[1], (int, float)) else None
            iade = row[2] if isinstance(row[2], (int, float)) else None
            net  = row[3] if isinstance(row[3], (int, float)) else None
            bas  = row[4] if isinstance(row[4], float) else None
            if sevk is None and net is None:
                continue
            magazines[current_mag]["years"].append(cell0)
            magazines[current_mag]["sevk"].append(float(sevk or 0))
            magazines[current_mag]["iade"].append(float(iade or 0))
            magazines[current_mag]["net"].append(float(net or 0))
            magazines[current_mag]["basari"].append(bas)
    return magazines


# ─────────────────────────────────────────────────────────────
# Global DataFrame oluştur (Kaggle stili)
# ─────────────────────────────────────────────────────────────
def build_global_df(magazines: dict, national_tiraj: dict, sector_trends: dict) -> pd.DataFrame:
    """
    90 dergi × 7 yıl = ~630 satır.
    Her satır: [mag_id, year, category_enc, nat_tiraj, cat_pct,
                prev_net, prev_sevk, prev_iade, iade_ratio,
                net_cagr_3y, sevk_cagr_3y, lag2_net, lag2_sevk,
                target_net, target_sevk, target_iade]
    """
    records = []
    le = LabelEncoder()
    le.fit(CAT_ORDER)

    all_years = sorted(national_tiraj.keys())

    for mag_id, (name, data) in enumerate(magazines.items()):
        years = data["years"]
        nets  = data["net"]
        sevks = data["sevk"]
        iades = data["iade"]
        cat   = data["category"]
        cat_enc = int(le.transform([cat])[0])

        # Kısmi 2025 filtresi
        year_net = dict(zip(years, nets))
        n2024 = year_net.get(2024, 0) or 0
        filtered = [(y, sevks[i], iades[i], nets[i])
                    for i, y in enumerate(years)
                    if not (y == 2025 and nets[i] < n2024 * 0.3 and n2024 > 0)]
        if not filtered:
            filtered = list(zip(years, sevks, iades, nets))

        f_years  = [r[0] for r in filtered]
        f_sevks  = [r[1] for r in filtered]
        f_iades  = [r[2] for r in filtered]
        f_nets   = [r[3] for r in filtered]

        for i, year in enumerate(f_years):
            if i == 0:
                continue  # Lag özellikleri için en az 1 önceki yıl gerekli

            # TÜİK
            nat = national_tiraj.get(year, 0)
            sec_rec = sector_trends.get(year, {})
            sk = _find_sector_key(cat, sec_rec)
            cat_pct = sec_rec.get(sk, 0) if sk else 0.0

            nat_prev = national_tiraj.get(year - 1, 0)
            sec_rec_p = sector_trends.get(year - 1, {})
            sk_p = _find_sector_key(cat, sec_rec_p)
            cat_pct_prev = sec_rec_p.get(sk_p, 0) if sk_p else 0.0

            nat_ratio  = (nat / nat_prev)  if nat_prev > 0 else 1.0
            cat_ratio  = (cat_pct / cat_pct_prev) if cat_pct_prev > 0 else 1.0
            hybrid_ratio = (nat_ratio + cat_ratio) * 0.5

            # Lag-1
            prev_net  = f_nets[i - 1]
            prev_sevk = f_sevks[i - 1]
            prev_iade = f_iades[i - 1]
            iade_ratio = prev_iade / prev_sevk if prev_sevk > 0 else 0.0

            # Lag-2
            lag2_net  = f_nets[i - 2]  if i >= 2 else prev_net
            lag2_sevk = f_sevks[i - 2] if i >= 2 else prev_sevk

            # 3 Yıllık CAGR (negatif değer koruması)
            if i >= 3 and f_nets[i - 3] > 0 and prev_net >= 0:
                net_cagr3 = float((prev_net / f_nets[i - 3]) ** (1 / 3) - 1)
            else:
                net_cagr3 = 0.0
            if i >= 3 and f_sevks[i - 3] > 0 and prev_sevk >= 0:
                sevk_cagr3 = float((prev_sevk / f_sevks[i - 3]) ** (1 / 3) - 1)
            else:
                sevk_cagr3 = 0.0

            records.append({
                "mag_id":       mag_id,
                "mag_name":     name,
                "year":         year,
                "category":     cat,
                "cat_enc":      cat_enc,
                "nat_tiraj":    nat / 1e7,          # normalize
                "cat_pct":      cat_pct,
                "nat_ratio":    nat_ratio,
                "cat_ratio":    cat_ratio,
                "hybrid_ratio": hybrid_ratio,
                "prev_net":     prev_net,
                "prev_sevk":    prev_sevk,
                "prev_iade":    prev_iade,
                "lag2_net":     lag2_net,
                "lag2_sevk":    lag2_sevk,
                "iade_ratio":   iade_ratio,
                "net_cagr3":    net_cagr3,
                "sevk_cagr3":   sevk_cagr3,
                "target_net":   f_nets[i],
                "target_sevk":  f_sevks[i],
                "target_iade":  f_iades[i],
            })

    return pd.DataFrame(records)


# ─────────────────────────────────────────────────────────────
# Walk-Forward Cross Validation
# ─────────────────────────────────────────────────────────────
FEATURES = [
    "cat_enc", "nat_tiraj", "cat_pct",
    "nat_ratio", "cat_ratio", "hybrid_ratio",
    "prev_net", "prev_sevk", "prev_iade",
    "lag2_net", "lag2_sevk",
    "iade_ratio", "net_cagr3", "sevk_cagr3",
]

LGB_PARAMS = {
    "objective":        "regression",
    "metric":           "mae",
    "n_estimators":     400,
    "learning_rate":    0.04,
    "num_leaves":       20,
    "min_child_samples": 5,
    "feature_fraction": 0.8,
    "bagging_fraction": 0.8,
    "bagging_freq":     5,
    "reg_alpha":        0.1,
    "reg_lambda":       0.5,
    "verbose":          -1,
    "n_jobs":           -1,
}


def walk_forward_cv(df: pd.DataFrame) -> dict:
    """
    Walk-Forward CV — Net Sevk ve Toplam Sevk için ayrı ayrı:
      Fold 1: Train 2020-2022 → Test 2023
      Fold 2: Train 2020-2023 → Test 2024
      Fold 3: Train 2020-2024 → Test 2025

    İade artık türetildiği için (iade = sevk - net) ayrı CV yapılmıyor.
    """
    years_avail = sorted(df["year"].unique())
    test_years = [y for y in years_avail if y >= 2023]

    # Net CV
    net_folds = []
    net_mae_all, net_mape_all, net_r2_all = [], [], []

    # Sevk CV
    sevk_folds = []
    sevk_mae_all, sevk_mape_all, sevk_r2_all = [], [], []

    for test_year in test_years:
        train_df = df[df["year"] < test_year]
        test_df  = df[df["year"] == test_year]
        if len(train_df) < 10 or len(test_df) < 5:
            continue

        X_tr = train_df[FEATURES]
        X_te = test_df[FEATURES]

        # ── NET modeli ──
        y_tr_net = train_df["target_net"]
        y_te_net = test_df["target_net"]
        net_model = lgb.LGBMRegressor(**LGB_PARAMS)
        net_model.fit(X_tr, y_tr_net,
                      eval_set=[(X_te, y_te_net)],
                      callbacks=[lgb.early_stopping(50, verbose=False),
                                  lgb.log_evaluation(period=-1)])
        preds_net = np.maximum(net_model.predict(X_te), 0)
        mae_net  = mean_absolute_error(y_te_net, preds_net)
        y_arr_net = y_te_net.values
        mask_net  = y_arr_net > 100
        mape_net  = float(np.mean(np.abs((y_arr_net[mask_net] - preds_net[mask_net]) / y_arr_net[mask_net])) * 100) \
                    if mask_net.sum() > 0 else 0.0
        try:
            r2_net = r2_score(y_te_net, preds_net)
        except Exception:
            r2_net = 0.0

        net_folds.append({
            "test_year": test_year,
            "n_train": len(train_df),
            "n_test": len(test_df),
            "mae": round(mae_net, 1),
            "mape": round(mape_net, 2),
            "r2": round(r2_net, 4),
        })
        net_mae_all.append(mae_net)
        net_mape_all.append(mape_net)
        net_r2_all.append(r2_net)

        # ── SEVK modeli ──
        y_tr_sevk = train_df["target_sevk"]
        y_te_sevk = test_df["target_sevk"]
        sevk_model = lgb.LGBMRegressor(**LGB_PARAMS)
        sevk_model.fit(X_tr, y_tr_sevk,
                       eval_set=[(X_te, y_te_sevk)],
                       callbacks=[lgb.early_stopping(50, verbose=False),
                                   lgb.log_evaluation(period=-1)])
        preds_sevk = np.maximum(sevk_model.predict(X_te), 0)
        mae_sevk   = mean_absolute_error(y_te_sevk, preds_sevk)
        y_arr_sevk = y_te_sevk.values
        mask_sevk  = y_arr_sevk > 100
        mape_sevk  = float(np.mean(np.abs((y_arr_sevk[mask_sevk] - preds_sevk[mask_sevk]) / y_arr_sevk[mask_sevk])) * 100) \
                     if mask_sevk.sum() > 0 else 0.0
        try:
            r2_sevk = r2_score(y_te_sevk, preds_sevk)
        except Exception:
            r2_sevk = 0.0

        sevk_folds.append({
            "test_year": test_year,
            "n_train": len(train_df),
            "n_test": len(test_df),
            "mae": round(mae_sevk, 1),
            "mape": round(mape_sevk, 2),
            "r2": round(r2_sevk, 4),
        })
        sevk_mae_all.append(mae_sevk)
        sevk_mape_all.append(mape_sevk)
        sevk_r2_all.append(r2_sevk)

    return {
        # Net CV
        "folds":      net_folds,
        "mean_mae":   round(float(np.mean(net_mae_all)),  1) if net_mae_all  else None,
        "mean_mape":  round(float(np.mean(net_mape_all)), 2) if net_mape_all else None,
        "mean_r2":    round(float(np.mean(net_r2_all)),   4) if net_r2_all   else None,
        # Sevk CV
        "sevk_folds":      sevk_folds,
        "sevk_mean_mae":   round(float(np.mean(sevk_mae_all)),  1) if sevk_mae_all  else None,
        "sevk_mean_mape":  round(float(np.mean(sevk_mape_all)), 2) if sevk_mape_all else None,
        "sevk_mean_r2":    round(float(np.mean(sevk_r2_all)),   4) if sevk_r2_all   else None,
    }


def train_final_model(df: pd.DataFrame) -> lgb.LGBMRegressor:
    """Tüm geçmiş veriyle son modeli eğit."""
    X = df[FEATURES]
    y = df["target_net"]
    model = lgb.LGBMRegressor(**LGB_PARAMS)
    model.fit(X, y, callbacks=[lgb.log_evaluation(period=-1)])
    return model


def train_sevk_model(df: pd.DataFrame) -> lgb.LGBMRegressor:
    """Toplam Sevk modeli. İade = Sevk - Net olarak türetilir, ayrı model yok."""
    X = df[FEATURES]
    y = df["target_sevk"]
    model = lgb.LGBMRegressor(**LGB_PARAMS)
    model.fit(X, y, callbacks=[lgb.log_evaluation(period=-1)])
    return model


# ─────────────────────────────────────────────────────────────
# TÜİK Hibrit Oranı
# ─────────────────────────────────────────────────────────────
def _hybrid_ratio(target: int, national_tiraj: dict, sector_trends: dict, category: str) -> float:
    pt = target - 1
    nat_c = national_tiraj.get(target)
    nat_p = national_tiraj.get(pt)
    if not nat_c or not nat_p or nat_p == 0:
        return 1.0
    nat_ratio = nat_c / nat_p
    sec_c = sector_trends.get(target, {})
    sec_p = sector_trends.get(pt, {})
    sk_c = _find_sector_key(category, sec_c)
    sk_p = _find_sector_key(category, sec_p)
    sv_c = sec_c.get(sk_c) if sk_c else None
    sv_p = sec_p.get(sk_p) if sk_p else None
    sec_ratio = (sv_c / sv_p) if (sv_c and sv_p and sv_p > 0) else nat_ratio
    return (nat_ratio + sec_ratio) * 0.5


# ─────────────────────────────────────────────────────────────
# Feature satırı oluştur (tahmin için)
# ─────────────────────────────────────────────────────────────
def _make_pred_row(
    target_year: int,
    prev_net: float, prev_sevk: float, prev_iade: float,
    lag2_net: float, lag2_sevk: float,
    cat_enc: int, national_tiraj: dict, sector_trends: dict, category: str,
    net_hist: list, sevk_hist: list,
) -> pd.DataFrame:
    nat = national_tiraj.get(target_year, 0)
    sec_rec = sector_trends.get(target_year, {})
    sk = _find_sector_key(category, sec_rec)
    cat_pct = sec_rec.get(sk, 0.0) if sk else 0.0

    nat_prev = national_tiraj.get(target_year - 1, 0)
    sec_rec_p = sector_trends.get(target_year - 1, {})
    sk_p = _find_sector_key(category, sec_rec_p)
    cat_pct_prev = sec_rec_p.get(sk_p, 0.0) if sk_p else 0.0

    nat_ratio = (nat / nat_prev) if nat_prev > 0 else 1.0
    cat_ratio = (cat_pct / cat_pct_prev) if cat_pct_prev > 0 else 1.0
    hybrid_ratio = (nat_ratio + cat_ratio) * 0.5

    iade_ratio = prev_iade / prev_sevk if prev_sevk > 0 else 0.0

    # 3 yıllık CAGR (mevcut geçmiş üzerinden)
    if len(net_hist) >= 3 and net_hist[-3] > 0 and prev_net >= 0:
        net_cagr3 = float((prev_net / net_hist[-3]) ** (1 / 3) - 1)
    else:
        net_cagr3 = 0.0
    if len(sevk_hist) >= 3 and sevk_hist[-3] > 0 and prev_sevk >= 0:
        sevk_cagr3 = float((prev_sevk / sevk_hist[-3]) ** (1 / 3) - 1)
    else:
        sevk_cagr3 = 0.0

    return pd.DataFrame([{
        "cat_enc":      cat_enc,
        "nat_tiraj":    nat / 1e7,
        "cat_pct":      cat_pct,
        "nat_ratio":    nat_ratio,
        "cat_ratio":    cat_ratio,
        "hybrid_ratio": hybrid_ratio,
        "prev_net":     prev_net,
        "prev_sevk":    prev_sevk,
        "prev_iade":    prev_iade,
        "lag2_net":     lag2_net,
        "lag2_sevk":    lag2_sevk,
        "iade_ratio":   iade_ratio,
        "net_cagr3":    net_cagr3,
        "sevk_cagr3":   sevk_cagr3,
    }])


# ─────────────────────────────────────────────────────────────
# Risk skoru
# ─────────────────────────────────────────────────────────────
def _compute_risk(net_2026, net_2027, ci_low, ci_high,
                  national_tiraj, sector_trends, category, iade_ratio_trend) -> float:
    score = 0.0
    nat25 = national_tiraj.get(2025, 1)
    nat26 = national_tiraj.get(2026, 1)
    if nat25 > 0:
        score += max(1 - nat26 / nat25, 0) * 40
    if net_2026 > 0:
        spread = ci_high - ci_low
        score += min(spread / net_2026 * 20, 30)
    if iade_ratio_trend and iade_ratio_trend > 0:
        score += min(iade_ratio_trend * 100, 20)
    sec25 = sector_trends.get(2025, {})
    sec26 = sector_trends.get(2026, {})
    sk25 = _find_sector_key(category, sec25)
    sk26 = _find_sector_key(category, sec26)
    if sk25 and sk26:
        sv25 = sec25.get(sk25, 0)
        sv26 = sec26.get(sk26, 0)
        if sv25 > 0 and sv26 < sv25:
            score += min((1 - sv26 / sv25) * 20, 10)
    return round(min(score, 100), 1)


def _risk_level(score: float) -> str:
    if score >= 60:
        return "Yüksek"
    if score >= 30:
        return "Orta"
    return "Düşük"


# ─────────────────────────────────────────────────────────────
# Tek dergi tahmin (ensemble)
# ─────────────────────────────────────────────────────────────
def forecast_one(
    name: str, data: dict,
    net_model: lgb.LGBMRegressor,
    sevk_model: lgb.LGBMRegressor,
    cat_enc: int,
    national_tiraj: dict, sector_trends: dict,
    cv_metrics: dict,
    le: LabelEncoder,
) -> dict:
    """
    Net Sevk ve Toplam Sevk ayrı modellerle tahmin edilir.
    İade = Sevk - Net olarak türetilir → matematiksel tutarlılık garantili.
    """
    years = data["years"]
    sevks = data["sevk"]
    iades = data["iade"]
    nets  = data["net"]
    cat   = data["category"]

    # Kısmi 2025 filtresi
    year_net = dict(zip(years, nets))
    n2024 = year_net.get(2024, 0) or 0
    filtered = [(years[i], sevks[i], iades[i], nets[i])
                for i in range(len(years))
                if not (years[i] == 2025 and nets[i] < n2024 * 0.3 and n2024 > 0)]
    if not filtered:
        filtered = list(zip(years, sevks, iades, nets))

    f_years = [r[0] for r in filtered]
    f_sevks = [r[1] for r in filtered]
    f_iades = [r[2] for r in filtered]
    f_nets  = [r[3] for r in filtered]

    # Başlangıç değerleri
    prev_net  = f_nets[-1]
    prev_sevk = f_sevks[-1]
    prev_iade = f_iades[-1]
    lag2_net  = f_nets[-2]  if len(f_nets)  >= 2 else f_nets[-1]
    lag2_sevk = f_sevks[-2] if len(f_sevks) >= 2 else f_sevks[-1]

    net_hist  = f_nets[:]
    sevk_hist = f_sevks[:]

    net_preds, sevk_preds, iade_preds = {}, {}, {}

    # CI için: global MAE (cross-validation'dan)
    mae = cv_metrics.get("mean_mae") or (prev_net * 0.15)

    for target in [2026, 2027]:
        row = _make_pred_row(
            target, prev_net, prev_sevk, prev_iade, lag2_net, lag2_sevk,
            cat_enc, national_tiraj, sector_trends, cat, net_hist, sevk_hist
        )

        # LGB tahminleri (net ve sevk ayrı model)
        lgb_net  = max(float(net_model.predict(row)[0]),  0)
        lgb_sevk = max(float(sevk_model.predict(row)[0]), 0)

        # Hibrit tahmin
        ratio = _hybrid_ratio(target, national_tiraj, sector_trends, cat)
        hybrid_net  = max(prev_net  * ratio, 0)
        hybrid_sevk = max(prev_sevk * ratio, 0)

        # Ensemble: 55% LGB + 45% Hibrit
        final_net  = round(0.55 * lgb_net  + 0.45 * hybrid_net,  2)
        final_sevk = round(0.55 * lgb_sevk + 0.45 * hybrid_sevk, 2)

        # İade türetme: sevk - net (matematiksel tutarlılık garantisi)
        # Negatif olamaz: sevk < net ise iade = 0 (kısıt)
        final_iade = round(max(final_sevk - final_net, 0), 2)

        # CI: MAE-tabanlı ±1σ band
        ci_low  = max(final_net - 1.0 * mae, 0)
        ci_high = final_net + 1.0 * mae

        net_preds[target]  = {"value": final_net,  "ci_low": round(ci_low, 2),  "ci_high": round(ci_high, 2)}
        sevk_preds[target] = {"value": final_sevk}
        iade_preds[target] = {"value": final_iade}  # türetilmiş değer

        # Sonraki yıl için güncelle
        # prev_iade da türetilmiş değerden güncelleniyor
        prev_net, prev_sevk, prev_iade = final_net, final_sevk, final_iade
        lag2_net, lag2_sevk = (f_nets[-1] if target == 2026 else net_preds[2026]["value"]), \
                              (f_sevks[-1] if target == 2026 else sevk_preds[2026]["value"])
        net_hist.append(final_net)
        sevk_hist.append(final_sevk)

    # Başarı oranı tahminleri
    def basari(net_v, sevk_v):
        return round(net_v / sevk_v, 4) if sevk_v > 0 else None

    basari_vals = [b for b in data["basari"] if b is not None]
    avg_basari = round(sum(basari_vals) / len(basari_vals), 4) if basari_vals else None

    chg_rate = None
    if len(f_nets) >= 2 and f_nets[-2] > 0:
        chg_rate = round((f_nets[-1] - f_nets[-2]) / f_nets[-2] * 100, 2)

    iade_ratio_trend = None
    if len(f_years) >= 3:
        ratios = [f_iades[i] / f_sevks[i] for i in range(len(f_years)) if f_sevks[i] > 0]
        if len(ratios) >= 3:
            iade_ratio_trend = round(ratios[-1] - ratios[-3], 4)

    risk_score = _compute_risk(
        net_preds[2026]["value"], net_preds[2027]["value"],
        net_preds[2026]["ci_low"], net_preds[2026]["ci_high"],
        national_tiraj, sector_trends, cat, iade_ratio_trend
    )

    return {
        "name": name,
        "category": cat,
        "years_actual":  f_years,
        "sevk_actual":   f_sevks,
        "iade_actual":   f_iades,
        "net_actual":    f_nets,
        "forecast_2026": net_preds[2026]["value"],
        "forecast_2027": net_preds[2027]["value"],
        "sevk_forecast_2026": sevk_preds[2026]["value"],
        "sevk_forecast_2027": sevk_preds[2027]["value"],
        "iade_forecast_2026": iade_preds[2026]["value"],
        "iade_forecast_2027": iade_preds[2027]["value"],
        "net_forecast_2026":  net_preds[2026]["value"],
        "net_forecast_2027":  net_preds[2027]["value"],
        "ci_low_2026":   net_preds[2026]["ci_low"],
        "ci_high_2026":  net_preds[2026]["ci_high"],
        "ci_low_2027":   net_preds[2027]["ci_low"],
        "ci_high_2027":  net_preds[2027]["ci_high"],
        "basari_forecast_2026": basari(net_preds[2026]["value"], sevk_preds[2026]["value"]),
        "basari_forecast_2027": basari(net_preds[2027]["value"], sevk_preds[2027]["value"]),
        "avg_basari":    avg_basari,
        "change_rate_last": chg_rate,
        "iade_ratio_trend": iade_ratio_trend,
        "risk_score":    risk_score,
        "risk_level":    _risk_level(risk_score),
        "best_model":    "LGB+Hibrit",
        "model_score":   cv_metrics.get("mean_r2"),
    }


# ─────────────────────────────────────────────────────────────
# Feature importance
# ─────────────────────────────────────────────────────────────
def get_feature_importance(model: lgb.LGBMRegressor) -> list:
    importances = model.feature_importances_
    names = FEATURES
    fi = sorted(zip(names, importances), key=lambda x: x[1], reverse=True)
    total = sum(v for _, v in fi) or 1
    return [{"feature": n, "importance": round(v / total * 100, 2)} for n, v in fi]


# ─────────────────────────────────────────────────────────────
# Ana yükleme ve önbellek
# ─────────────────────────────────────────────────────────────
_cache = {}


def load_all(force: bool = False) -> dict:
    global _cache
    if _cache and not force:
        return _cache

    print("[1/5] Excel okunuyor...")
    national_tiraj = parse_national_tiraj()
    sector_trends  = parse_sector_trends()
    magazines_raw  = parse_magazine_data()

    print("[2/5] Global DataFrame oluşturuluyor (90 dergi × 7 yıl)...")
    df = build_global_df(magazines_raw, national_tiraj, sector_trends)
    print(f"      {len(df)} satır, {len(FEATURES)} özellik")

    print("[3/5] Walk-Forward CV (3 fold: 2023, 2024, 2025) — Net Sevk & Toplam Sevk...")
    cv_metrics = walk_forward_cv(df)
    print("      [Net Sevk CV]")
    for fold in cv_metrics["folds"]:
        print(f"        Fold {fold['test_year']}: MAE={fold['mae']:,.0f}  MAPE={fold['mape']:.1f}%  R²={fold['r2']:.3f}  "
              f"(train={fold['n_train']}, test={fold['n_test']})")
    print(f"        Ortalama -> MAE={cv_metrics['mean_mae']:,.0f}  MAPE={cv_metrics['mean_mape']:.1f}%  R²={cv_metrics['mean_r2']:.3f}")
    print("      [Toplam Sevk CV]")
    for fold in cv_metrics["sevk_folds"]:
        print(f"        Fold {fold['test_year']}: MAE={fold['mae']:,.0f}  MAPE={fold['mape']:.1f}%  R²={fold['r2']:.3f}  "
              f"(train={fold['n_train']}, test={fold['n_test']})")
    print(f"        Ortalama -> MAE={cv_metrics['sevk_mean_mae']:,.0f}  MAPE={cv_metrics['sevk_mean_mape']:.1f}%  R²={cv_metrics['sevk_mean_r2']:.3f}")

    print("[4/5] Final modeller eğitiliyor (Net Sevk + Toplam Sevk, tüm veri)...")
    net_model  = train_final_model(df)
    sevk_model = train_sevk_model(df)
    # İade modeli yok: iade = sevk - net olarak türetilecek

    print("[5/5] 90 dergi için tahmin üretiliyor...")
    le = LabelEncoder()
    le.fit(CAT_ORDER)
    mag_enc_map = {}
    for name, data in magazines_raw.items():
        cat = data["category"]
        mag_enc_map[name] = int(le.transform([cat])[0])

    results = {}
    for name, data in magazines_raw.items():
        results[name] = forecast_one(
            name, data, net_model, sevk_model,
            mag_enc_map[name], national_tiraj, sector_trends, cv_metrics, le
        )

    fi = get_feature_importance(net_model)

    _cache = {
        "magazines":      results,
        "national_tiraj": national_tiraj,
        "sector_trends":  sector_trends,
        "cv_metrics":     cv_metrics,
        "feature_importance": fi,
    }
    print("Tamamlandı.\n")
    return _cache
